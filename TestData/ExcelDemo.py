import openpyxl

book = openpyxl.load_workbook("C:\\Users\\thod_\\Downloads\\PythonDemo.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=2)
dicti = {}
print(cell.value)

print(sheet.max_row)
print(sheet.max_column)
print(sheet['A2'].value)

for row in range(1, sheet.max_row+1):
    if sheet.cell(row=row, column=1).value == 'Testcase2':
        for column in range(2, sheet.max_column+1):
            dicti[sheet.cell(row=1, column=column).value] = sheet.cell(row=row, column=column).value


print(dicti)


