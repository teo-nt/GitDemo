import openpyxl


class HomePageData:

    test_homepage_data = [{'firstname': 'Rahul', 'lastname': 'Shetty', 'password': '528dffs', 'genderindex': 0},
                            {'firstname': 'tho', 'lastname': 'kihotis', 'password': 'bigbilly', 'genderindex': 0}]

    @staticmethod
    def get_test_data(test_case_name):
        dicti = {}
        book = openpyxl.load_workbook("C:\\Users\\thod_\\Downloads\\PythonDemo.xlsx")
        sheet = book.active
        for row in range(1, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == test_case_name:
                for column in range(2, sheet.max_column + 1):
                    dicti[sheet.cell(row=1, column=column).value] = sheet.cell(row=row, column=column).value

        return [dicti]